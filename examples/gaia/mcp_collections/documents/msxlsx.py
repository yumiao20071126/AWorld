import json
import os
import subprocess
import sys
import time
import traceback
import zipfile
from pathlib import Path
from typing import Any, Literal

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from pydantic import Field
from pydantic.fields import FieldInfo

from aworld.logs.util import Color
from examples.gaia.mcp_collections.base import ActionArguments, ActionCollection, ActionResponse
from examples.gaia.mcp_collections.documents.models import DocumentMetadata


class XLSXExtractionCollection(ActionCollection):
    """MCP service for Excel document content extraction using xlrd and pandas.

    Supports extraction from XLSX and XLS files.
    Provides LLM-friendly text output with structured metadata and media file handling.
    Extracts worksheets, formulas, charts, and embedded images.
    Includes screenshot functionality for visual representation of Excel data.
    """

    def __init__(self, arguments: ActionArguments) -> None:
        super().__init__(arguments)
        self._media_output_dir = self.workspace / "extracted_media"
        self._media_output_dir.mkdir(exist_ok=True)

        # Create screenshots directory
        self._screenshots_dir = self.workspace / "excel_screenshots"
        self._screenshots_dir.mkdir(exist_ok=True)

        self.supported_extensions: set = {
            ".xlsx",
            ".xls",
        }

        self._color_log("Excel Extraction Service initialized", Color.green, "debug")
        self._color_log(f"Media output directory: {self._media_output_dir}", Color.blue, "debug")
        self._color_log(f"Screenshots directory: {self._screenshots_dir}", Color.blue, "debug")

    def _create_excel_screenshot(self, file_path: Path, sheet_name: str = None) -> str:
        """Create a JPEG screenshot of the valid Excel area using pyautogui.

        Args:
            file_path: Path to the Excel file
            sheet_name: Specific sheet to screenshot (None for first sheet)

        Returns:
            Path to the generated JPEG screenshot
        """
        try:
            import pyautogui

            # Generate unique filename
            timestamp = int(time.time())
            screenshot_filename = f"{file_path.stem}_{sheet_name or 'sheet'}_{timestamp}.jpg"
            screenshot_path = self._screenshots_dir / screenshot_filename

            # Open Excel file with default application
            if sys.platform == "darwin":  # macOS
                subprocess.run(["open", str(file_path)], check=True)
            elif sys.platform == "win32":  # Windows
                subprocess.run(["start", str(file_path)], shell=True, check=True)
            else:  # Linux
                subprocess.run(["xdg-open", str(file_path)], check=True)

            # Wait for Excel to open
            time.sleep(3)

            # Take screenshot of the entire screen
            screenshot = pyautogui.screenshot()

            # Convert RGBA to RGB before saving as JPEG
            if screenshot.mode == "RGBA":
                screenshot = screenshot.convert("RGB")

            screenshot.save(screenshot_path, "JPEG", quality=95)

            self._color_log(f"Created Excel screenshot: {screenshot_filename}", Color.green)
            return str(screenshot_path)

        except Exception as e:
            self.logger.error(f"Failed to create Excel screenshot with pyautogui: {str(e)}")
            raise

    def _extract_embedded_media_xlsx(self, file_path: Path) -> list[dict[str, str]]:
        """Extract embedded media from XLSX files.

        Args:
            file_path: Path to the XLSX file

        Returns:
            List of dictionaries containing media information
        """
        saved_media = []

        try:
            # Load workbook to extract images
            workbook = load_workbook(file_path, data_only=False)

            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]

                # Extract images from worksheet
                if hasattr(worksheet, "_images"):
                    for idx, image in enumerate(worksheet._images):
                        try:
                            # Generate unique filename
                            image_filename = f"{file_path.stem}_{sheet_name}_img_{idx}.png"
                            image_path = self._media_output_dir / image_filename

                            # Save image
                            if hasattr(image, "ref"):
                                # Extract image data
                                img_data = image._data()
                                if img_data:
                                    with open(image_path, "wb") as f:
                                        f.write(img_data)

                                    saved_media.append(
                                        {
                                            "type": "image",
                                            "path": str(image_path),
                                            "sheet": sheet_name,
                                            "filename": image_filename,
                                        }
                                    )

                                    self._color_log(f"Saved image: {image_filename}", Color.blue)
                        except Exception as e:
                            self.logger.warning(f"Failed to extract image {idx} from sheet {sheet_name}: {str(e)}")

            # Also try to extract from ZIP structure for additional media
            with zipfile.ZipFile(file_path, "r") as zip_file:
                media_files = [f for f in zip_file.namelist() if f.startswith("xl/media/")]

                for media_file in media_files:
                    try:
                        media_data = zip_file.read(media_file)
                        media_filename = f"{file_path.stem}_{Path(media_file).name}"
                        media_path = self._media_output_dir / media_filename

                        with open(media_path, "wb") as f:
                            f.write(media_data)

                        # Determine media type based on extension
                        media_ext = Path(media_file).suffix.lower()
                        if media_ext in [".png", ".jpg", ".jpeg", ".gif", ".bmp"]:
                            media_type = "image"
                        elif media_ext in [".mp3", ".wav", ".m4a"]:
                            media_type = "audio"
                        elif media_ext in [".mp4", ".avi", ".mov"]:
                            media_type = "video"
                        else:
                            media_type = "other"

                        saved_media.append(
                            {
                                "type": media_type,
                                "path": str(media_path),
                                "filename": media_filename,
                                "original_path": media_file,
                            }
                        )

                        self._color_log(f"Saved media: {media_filename}", Color.blue)
                    except Exception as e:
                        self.logger.warning(f"Failed to extract media {media_file}: {str(e)}")

        except Exception as e:
            self.logger.warning(f"Failed to extract media from XLSX: {str(e)}")

        return saved_media

    def _extract_excel_content(self, file_path: Path, sheet_names: list[str] | None = None) -> dict[str, Any]:
        """Extract content from Excel files using pandas and xlrd.

        Args:
            file_path: Path to the Excel file
            sheet_names: Specific sheets to process (None for all sheets)

        Returns:
            Dictionary containing extracted content and metadata
        """
        start_time = time.time()

        try:
            # Determine file type and read accordingly
            if file_path.suffix.lower() == ".xlsx":
                # Use openpyxl engine for XLSX files
                excel_file = pd.ExcelFile(file_path, engine="openpyxl")
            else:
                # Use xlrd engine for XLS files
                excel_file = pd.ExcelFile(file_path, engine="xlrd")

            # Get all sheet names if not specified
            if sheet_names is None:
                sheet_names = excel_file.sheet_names

            sheets_data = {}
            total_rows = 0
            total_cols = 0

            # Extract data from each sheet
            for sheet_name in sheet_names:
                if sheet_name in excel_file.sheet_names:
                    try:
                        # Read sheet data
                        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)

                        # Remove completely empty rows and columns
                        df = df.dropna(how="all").dropna(axis=1, how="all")

                        if not df.empty:
                            sheets_data[sheet_name] = {
                                "data": df,
                                "shape": df.shape,
                                "columns": df.columns.tolist(),
                                "non_empty_cells": df.count().sum(),
                            }

                            total_rows += df.shape[0]
                            total_cols = max(total_cols, df.shape[1])
                        else:
                            sheets_data[sheet_name] = {"data": df, "shape": (0, 0), "columns": [], "non_empty_cells": 0}

                    except Exception as e:
                        self.logger.warning(f"Failed to read sheet '{sheet_name}': {str(e)}")
                        sheets_data[sheet_name] = {
                            "error": str(e),
                            "shape": (0, 0),
                            "columns": [],
                            "non_empty_cells": 0,
                        }

            processing_time = time.time() - start_time

            return {
                "sheets_data": sheets_data,
                "sheet_names": list(sheets_data.keys()),
                "total_sheets": len(sheets_data),
                "total_rows": total_rows,
                "total_columns": total_cols,
                "processing_time": processing_time,
                "file_engine": "openpyxl" if file_path.suffix.lower() == ".xlsx" else "xlrd",
            }

        except Exception as e:
            self.logger.error(f"Failed to extract Excel content: {str(e)}")
            raise

    def _format_content_for_llm(
        self, extraction_result: dict[str, Any], output_format: str, include_empty_cells: bool = False
    ) -> str:
        """Format extracted Excel content to be LLM-friendly.

        Args:
            extraction_result: Result from _extract_excel_content
            output_format: Desired output format
            include_empty_cells: Whether to include empty cells in output

        Returns:
            Formatted content string
        """
        sheets_data = extraction_result["sheets_data"]

        if output_format.lower() == "markdown":
            content_parts = []
            content_parts.append("# Excel Document Content\n")
            content_parts.append(f"**Total Sheets:** {extraction_result['total_sheets']}\n")
            content_parts.append(f"**Processing Engine:** {extraction_result['file_engine']}\n\n")

            for sheet_name, sheet_info in sheets_data.items():
                content_parts.append(f"## Sheet: {sheet_name}\n")

                if "error" in sheet_info:
                    content_parts.append(f"**Error:** {sheet_info['error']}\n\n")
                    continue

                df: pd.DataFrame = sheet_info["data"]
                shape = sheet_info["shape"]

                content_parts.append(f"**Dimensions:** {shape[0]} rows × {shape[1]} columns\n")
                content_parts.append(f"**Non-empty cells:** {sheet_info['non_empty_cells']}\n\n")

                if not df.empty:
                    # Convert DataFrame to markdown table
                    if include_empty_cells:
                        # Fill NaN values with empty string for display
                        df_display = df.fillna("")
                    else:
                        # Keep NaN values as they are
                        df_display = df

                    # Convert to markdown table
                    try:
                        markdown_table = df_display.to_markdown(index=False, tablefmt="pipe")
                        content_parts.append(f"### Data:\n{markdown_table}\n\n")
                    except Exception:
                        # Fallback to string representation
                        content_parts.append(f"### Data (text format):\n```\n{df_display.to_string()}\n```\n\n")
                else:
                    content_parts.append("*Sheet is empty*\n\n")

            return "".join(content_parts)

        elif output_format.lower() == "json":
            json_data = {
                "document_info": {
                    "total_sheets": extraction_result["total_sheets"],
                    "total_rows": extraction_result["total_rows"],
                    "total_columns": extraction_result["total_columns"],
                    "processing_engine": extraction_result["file_engine"],
                },
                "sheets": {},
            }

            for sheet_name, sheet_info in sheets_data.items():
                if "error" in sheet_info:
                    json_data["sheets"][sheet_name] = {"error": sheet_info["error"], "shape": sheet_info["shape"]}
                    continue

                df = sheet_info["data"]

                if not df.empty:
                    # Convert DataFrame to records
                    if include_empty_cells:
                        df_records = df.fillna("").to_dict("records")
                    else:
                        df_records = df.to_dict("records")

                    json_data["sheets"][sheet_name] = {
                        "shape": sheet_info["shape"],
                        "non_empty_cells": sheet_info["non_empty_cells"],
                        "data": df_records,
                    }
                else:
                    json_data["sheets"][sheet_name] = {"shape": sheet_info["shape"], "non_empty_cells": 0, "data": []}

            return json.dumps(json_data, indent=2, default=str)

        elif output_format.lower() == "html":
            html_parts = []
            html_parts.append("<html><body>")
            html_parts.append("<h1>Excel Document Content</h1>")
            html_parts.append(f"<p><strong>Total Sheets:</strong> {extraction_result['total_sheets']}</p>")
            html_parts.append(f"<p><strong>Processing Engine:</strong> {extraction_result['file_engine']}</p>")

            for sheet_name, sheet_info in sheets_data.items():
                html_parts.append(f"<h2>Sheet: {sheet_name}</h2>")

                if "error" in sheet_info:
                    html_parts.append(f"<p><strong>Error:</strong> {sheet_info['error']}</p>")
                    continue

                df = sheet_info["data"]
                shape = sheet_info["shape"]

                html_parts.append(f"<p><strong>Dimensions:</strong> {shape[0]} rows × {shape[1]} columns</p>")
                html_parts.append(f"<p><strong>Non-empty cells:</strong> {sheet_info['non_empty_cells']}</p>")

                if not df.empty:
                    # Convert DataFrame to HTML table
                    if include_empty_cells:
                        df_display = df.fillna("")
                    else:
                        df_display = df

                    html_table = df_display.to_html(index=False, escape=False, table_id=f"sheet_{sheet_name}")
                    html_parts.append(html_table)
                else:
                    html_parts.append("<p><em>Sheet is empty</em></p>")

            html_parts.append("</body></html>")
            return "".join(html_parts)

        else:  # text format
            content_parts = []
            content_parts.append(f"Excel Document Content\n{'=' * 50}\n")
            content_parts.append(f"Total Sheets: {extraction_result['total_sheets']}\n")
            content_parts.append(f"Processing Engine: {extraction_result['file_engine']}\n\n")

            for sheet_name, sheet_info in sheets_data.items():
                content_parts.append(f"Sheet: {sheet_name}\n{'-' * 30}\n")

                if "error" in sheet_info:
                    content_parts.append(f"Error: {sheet_info['error']}\n\n")
                    continue

                df = sheet_info["data"]
                shape = sheet_info["shape"]

                content_parts.append(f"Dimensions: {shape[0]} rows × {shape[1]} columns\n")
                content_parts.append(f"Non-empty cells: {sheet_info['non_empty_cells']}\n\n")

                if not df.empty:
                    if include_empty_cells:
                        df_display = df.fillna("")
                    else:
                        df_display = df

                    content_parts.append(f"Data:\n{df_display.to_string()}\n\n")
                else:
                    content_parts.append("Sheet is empty\n\n")

            return "".join(content_parts)

    def mcp_extract_excel_content(
        self,
        file_path: str = Field(description="Path to the Excel document file to extract content from"),
        output_format: Literal["markdown", "json", "html", "text"] = Field(
            default="markdown", description="Output format: 'markdown', 'json', 'html', or 'text'"
        ),
        extract_images: bool = Field(default=True, description="Whether to extract and save images from the document"),
        create_screenshot: bool = Field(
            default=False, description="Whether to create a JPEG screenshot of the Excel data"
        ),
        sheet_names: str | None = Field(
            default=None, description="Comma-separated list of specific sheet names to process (None for all sheets)"
        ),
        include_empty_cells: bool = Field(default=False, description="Whether to include empty cells in the output"),
        screenshot_max_rows: int = Field(default=50, description="Maximum rows to include in screenshot"),
        screenshot_max_cols: int = Field(default=20, description="Maximum columns to include in screenshot"),
    ) -> ActionResponse:
        """Extract content from Excel documents using pandas and xlrd.

        This tool provides comprehensive Excel document content extraction with support for:
        - XLSX and XLS files
        - Multiple worksheets
        - Text and numeric data extraction
        - Image and media extraction (XLSX only)
        - JPEG screenshot generation of Excel data
        - Metadata collection
        - LLM-optimized output formatting

        Args:
            file_path: Path to the Excel file
            output_format: Desired output format
            extract_images: Whether to extract embedded images
            create_screenshot: Whether to create a JPEG screenshot
            sheet_names: Specific sheets to process
            include_empty_cells: Whether to include empty cells
            screenshot_max_rows: Maximum rows in screenshot
            screenshot_max_cols: Maximum columns in screenshot

        Returns:
            ActionResponse with extracted content, metadata, media file paths, and screenshot path
        """
        try:
            # Handle FieldInfo objects
            if isinstance(file_path, FieldInfo):
                file_path = file_path.default
            if isinstance(output_format, FieldInfo):
                output_format = output_format.default
            if isinstance(extract_images, FieldInfo):
                extract_images = extract_images.default
            if isinstance(create_screenshot, FieldInfo):
                create_screenshot = create_screenshot.default
            if isinstance(sheet_names, FieldInfo):
                sheet_names = sheet_names.default
            if isinstance(include_empty_cells, FieldInfo):
                include_empty_cells = include_empty_cells.default
            if isinstance(screenshot_max_rows, FieldInfo):
                screenshot_max_rows = screenshot_max_rows.default
            if isinstance(screenshot_max_cols, FieldInfo):
                screenshot_max_cols = screenshot_max_cols.default

            # Validate input file
            file_path: Path = self._validate_file_path(file_path)
            self._color_log(f"Processing Excel document: {file_path.name}", Color.cyan)

            # Parse sheet names if provided
            target_sheets = None
            if sheet_names:
                target_sheets = [name.strip() for name in sheet_names.split(",")]

            # Extract content from Excel file
            extraction_result = self._extract_excel_content(file_path, target_sheets)

            # Extract embedded media if requested (XLSX only)
            saved_media = []
            if extract_images and file_path.suffix.lower() == ".xlsx":
                saved_media = self._extract_embedded_media_xlsx(file_path)
            elif extract_images and file_path.suffix.lower() == ".xls":
                self._color_log("Image extraction not supported for XLS files", Color.yellow)

            # Create screenshot if requested
            screenshot_path = None
            if create_screenshot:
                target_sheet = target_sheets[0] if target_sheets else None
                screenshot_path = self._create_excel_screenshot(file_path, target_sheet)

            # Format content for LLM consumption
            formatted_content = self._format_content_for_llm(extraction_result, output_format, include_empty_cells)

            # Prepare metadata
            file_stats = file_path.stat()

            # Create Excel-specific metadata
            excel_metadata = {
                "sheet_count": extraction_result["total_sheets"],
                "sheet_names": extraction_result["sheet_names"],
                "total_rows": extraction_result["total_rows"],
                "total_columns": extraction_result["total_columns"],
                "processing_engine": extraction_result["file_engine"],
                "extracted_images": [media["path"] for media in saved_media if media["type"] == "image"],
                "extracted_media": saved_media,
                "screenshot_path": screenshot_path,
                "include_empty_cells": include_empty_cells,
                "processed_sheets": target_sheets or extraction_result["sheet_names"],
            }

            document_metadata = DocumentMetadata(
                file_name=file_path.name,
                file_size=file_stats.st_size,
                file_type=file_path.suffix.lower(),
                absolute_path=str(file_path.absolute()),
                page_count=extraction_result["total_sheets"],  # Use sheet count as "page" count
                processing_time=extraction_result["processing_time"],
                extracted_images=[media["path"] for media in saved_media if media["type"] == "image"],
                extracted_media=saved_media,
                output_format=output_format,
                llm_enhanced=False,
                ocr_applied=False,
            )

            # Combine standard and Excel-specific metadata
            combined_metadata = document_metadata.model_dump()
            combined_metadata.update(excel_metadata)

            success_message = (
                f"Successfully extracted content from {file_path.name} "
                f"({len(formatted_content)} characters, {extraction_result['total_sheets']} sheets, "
                f"{len(saved_media)} media files"
            )

            if screenshot_path:
                success_message += f", screenshot saved to: {screenshot_path}"

            self._color_log(success_message, Color.green)

            return ActionResponse(success=True, message=formatted_content, metadata=combined_metadata)

        except FileNotFoundError as e:
            self.logger.error(f"File not found: {str(e)}: {traceback.format_exc()}")
            return ActionResponse(
                success=False, message=f"File not found: {str(e)}", metadata={"error_type": "file_not_found"}
            )
        except ValueError as e:
            self.logger.error(f"Invalid input: {str(e)}: {traceback.format_exc()}")
            return ActionResponse(
                success=False,
                message=f"Invalid input: {str(e)}",
                metadata={"error_type": "invalid_input"},
            )
        except Exception as e:
            self.logger.error(f"Excel extraction failed: {str(e)}: {traceback.format_exc()}")
            return ActionResponse(
                success=False,
                message=f"Excel extraction failed: {str(e)}",
                metadata={"error_type": "extraction_error"},
            )

    def mcp_create_excel_screenshot(
        self,
        file_path: str = Field(description="Path to the Excel document file"),
        sheet_name: str | None = Field(
            default=None, description="Specific sheet name to screenshot (None for first sheet)"
        ),
        max_rows: int = Field(default=50, description="Maximum number of rows to include"),
        max_cols: int = Field(default=20, description="Maximum number of columns to include"),
    ) -> ActionResponse:
        """Create a JPEG screenshot of the valid Excel area.

        This tool creates a visual representation of Excel data as a JPEG image,
        useful for further image processing or visual analysis.

        Args:
            file_path: Path to the Excel file
            sheet_name: Specific sheet to screenshot
            max_rows: Maximum rows to include in screenshot
            max_cols: Maximum columns to include in screenshot

        Returns:
            ActionResponse with screenshot file path and metadata
        """
        try:
            # Handle FieldInfo objects
            if isinstance(file_path, FieldInfo):
                file_path = file_path.default
            if isinstance(sheet_name, FieldInfo):
                sheet_name = sheet_name.default
            if isinstance(max_rows, FieldInfo):
                max_rows = max_rows.default
            if isinstance(max_cols, FieldInfo):
                max_cols = max_cols.default

            # Validate input file
            file_path: Path = self._validate_file_path(file_path)
            self._color_log(f"Creating screenshot for Excel document: {file_path.name}", Color.cyan)

            # Create screenshot
            screenshot_path = self._create_excel_screenshot(file_path, sheet_name)

            # Prepare metadata
            file_stats = file_path.stat()
            screenshot_stats = Path(screenshot_path).stat()

            metadata = {
                "source_file": str(file_path.absolute()),
                "source_file_size": file_stats.st_size,
                "screenshot_path": screenshot_path,
                "screenshot_size": screenshot_stats.st_size,
                "sheet_name": sheet_name,
                "max_rows_displayed": max_rows,
                "max_cols_displayed": max_cols,
                "format": "JPEG",
            }

            return ActionResponse(
                success=True,
                message=f"Excel screenshot created successfully. File saved to: {screenshot_path}",
                metadata=metadata,
            )

        except Exception as e:
            self.logger.error(f"Screenshot creation failed: {str(e)}: {traceback.format_exc()}")
            return ActionResponse(
                success=False,
                message=f"Screenshot creation failed: {str(e)}",
                metadata={"error_type": "screenshot_error"},
            )

    def mcp_list_supported_formats(self) -> ActionResponse:
        """List all supported Excel formats for extraction.

        Returns:
            ActionResponse with list of supported file formats and their descriptions
        """
        supported_formats = {
            "XLSX": "Excel 2007+ format files (.xlsx) - Full support including images",
            "XLS": "Excel 97-2003 format files (.xls) - Text and data only",
        }

        format_list = "\n".join(
            [f"**{format_name}**: {description}" for format_name, description in supported_formats.items()]
        )

        return ActionResponse(
            success=True,
            message=f"Supported Excel formats:\n\n{format_list}",
            metadata={"supported_formats": list(supported_formats.keys()), "total_formats": len(supported_formats)},
        )


# Example usage and entry point
if __name__ == "__main__":
    load_dotenv()

    # Default arguments for testing
    args = ActionArguments(
        name="excel_extraction_service",
        transport="stdio",
        workspace=os.getenv("AWORLD_WORKSPACE", "/tmp"),
    )

    # Initialize and run the Excel extraction service
    try:
        service = XLSXExtractionCollection(args)
        service.run()
    except Exception as e:
        print(f"An error occurred: {e}: {traceback.format_exc()}")
